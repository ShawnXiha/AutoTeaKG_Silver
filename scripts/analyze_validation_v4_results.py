import csv
import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
VAL_DIR = ROOT / "writing_outputs" / "20260412_autoteakg_silver_paper" / "data" / "validation_v4"
WORKBOOK = VAL_DIR / "AutoTeaKG_validation_v4_for_undergraduates.xlsx"
OUT_REPORT = ROOT / "writing_outputs" / "20260412_autoteakg_silver_paper" / "reviews" / "VALIDATION_RESULTS_v4.md"
OUT_JSON = VAL_DIR / "validation_results_v4.json"
OUT_CSV = VAL_DIR / "validation_results_by_field_v4.csv"
OUT_CLEAN = VAL_DIR / "validation_worksheet_v4_filled_cleaned.csv"
OUT_SNIPPET = ROOT / "writing_outputs" / "20260412_autoteakg_silver_paper" / "drafts" / "v4_snippets" / "validation_results_table.tex"


VALID_DECISIONS = {"correct", "minor_issue", "major_issue", "not_applicable"}
VALID_OVERALL = {"correct", "minor_issue", "major_issue", "not_applicable", "accept", "accept_with_correction", "reject"}

FIELDS = [
    ("activity_category_decision", "Activity category"),
    ("study_type_decision", "Study type"),
    ("evidence_level_decision", "Evidence level"),
    ("component_group_decision", "Component group"),
    ("processing_step_decision", "Processing step"),
    ("extraction_method_decision", "Extraction method"),
    ("mechanism_label_decision", "Mechanism label"),
]


def load_rows():
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb.worksheets[1]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        if any((row.get(k) not in (None, "") for k in headers if k.endswith("_decision") or k in ("overall_record_decision", "major_issue_type", "comments"))):
            rows.append(row)
    return rows


def normalize(value):
    if value is None:
        return ""
    return str(value).strip()


def clean_rows(rows):
    cleaned = []
    anomalies = []
    for row in rows:
        new_row = {k: normalize(v) for k, v in row.items()}
        for field, _ in FIELDS:
            value = new_row.get(field, "")
            if value and value not in VALID_DECISIONS:
                anomalies.append(
                    {
                        "audit_id": new_row.get("audit_id", ""),
                        "record_id": new_row.get("record_id", ""),
                        "field": field,
                        "value": value,
                        "action": "treated_as_blank_invalid_decision",
                    }
                )
                new_row[field] = ""
        overall = new_row.get("overall_record_decision", "")
        if overall and overall not in VALID_OVERALL:
            anomalies.append(
                {
                    "audit_id": new_row.get("audit_id", ""),
                    "record_id": new_row.get("record_id", ""),
                    "field": "overall_record_decision",
                    "value": overall,
                    "action": "treated_as_blank_invalid_overall",
                }
            )
            new_row["overall_record_decision"] = ""
        cleaned.append(new_row)
    return cleaned, anomalies


def summarize(rows):
    overall_counter = Counter(row.get("overall_record_decision", "") for row in rows if row.get("overall_record_decision", ""))
    issue_counter = Counter(row.get("major_issue_type", "") for row in rows if row.get("major_issue_type", ""))
    uncertainty_counter = Counter(row.get("uncertainty_class", "") for row in rows if row.get("uncertainty_class", ""))
    field_rows = []
    for field, label in FIELDS:
        nonblank = [row.get(field, "") for row in rows if row.get(field, "")]
        counter = Counter(nonblank)
        total = sum(counter.values())
        field_rows.append(
            {
                "field": label,
                "field_key": field,
                "completed": total,
                "correct": counter.get("correct", 0),
                "minor_issue": counter.get("minor_issue", 0),
                "major_issue": counter.get("major_issue", 0),
                "not_applicable": counter.get("not_applicable", 0),
                "completion_rate_vs_47": round(total / len(rows), 3) if rows else 0,
                "correct_rate_on_completed": round(counter.get("correct", 0) / total, 3) if total else 0,
                "acceptable_rate_on_completed": round((counter.get("correct", 0) + counter.get("minor_issue", 0)) / total, 3) if total else 0,
            }
        )
    return {
        "reviewed_records": len(rows),
        "overall_counter": overall_counter,
        "major_issue_counter": issue_counter,
        "uncertainty_counter": uncertainty_counter,
        "field_rows": field_rows,
    }


def write_outputs(cleaned_rows, anomalies, summary):
    with OUT_CLEAN.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(cleaned_rows[0].keys()))
        writer.writeheader()
        writer.writerows(cleaned_rows)

    with OUT_CSV.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(summary["field_rows"][0].keys()))
        writer.writeheader()
        writer.writerows(summary["field_rows"])

    payload = {
        "reviewed_records": summary["reviewed_records"],
        "overall_counter": dict(summary["overall_counter"]),
        "major_issue_counter": dict(summary["major_issue_counter"]),
        "uncertainty_counter": dict(summary["uncertainty_counter"]),
        "field_rows": summary["field_rows"],
        "anomalies": anomalies,
    }
    OUT_JSON.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")

    lines = [
        "# Validation Results v4",
        "",
        f"- Reviewed records: {summary['reviewed_records']} / 48",
        f"- Overall decisions: {dict(summary['overall_counter'])}",
        f"- Major issue types: {dict(summary['major_issue_counter'])}",
        f"- Uncertainty distribution in reviewed sample: {dict(summary['uncertainty_counter'])}",
        "",
        "## Field-Level Results",
        "",
        "| Field | Completed | Correct | Minor | Major | N/A | Correct rate | Acceptable rate |",
        "|---|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for row in summary["field_rows"]:
        lines.append(
            f"| {row['field']} | {row['completed']} | {row['correct']} | {row['minor_issue']} | {row['major_issue']} | {row['not_applicable']} | {row['correct_rate_on_completed']:.1%} | {row['acceptable_rate_on_completed']:.1%} |"
        )
    if anomalies:
        lines.extend(["", "## Annotation anomalies cleaned", ""])
        for a in anomalies:
            lines.append(f"- {a['audit_id']} / {a['record_id']} / {a['field']}: `{a['value']}` -> blank")
    OUT_REPORT.write_text("\n".join(lines), encoding="utf-8")

    tex_lines = [
        r"\begin{table}[t]",
        r"\centering",
        r"\small",
        r"\caption{External validation results on the stratified 47-record sample completed by human reviewers. Acceptable rate = correct + minor issue over completed judgments.}",
        r"\label{tab:validation_results}",
        r"\begin{tabular}{lrrrrrr}",
        r"\toprule",
        r"Field & N & Correct & Minor & Major & N/A & Acceptable \\",
        r"\midrule",
    ]
    for row in summary["field_rows"]:
        tex_lines.append(
            f"{row['field']} & {row['completed']} & {row['correct']} & {row['minor_issue']} & {row['major_issue']} & {row['not_applicable']} & {row['acceptable_rate_on_completed']:.0%}".replace('%', r'\%') + r" \\"
        )
    tex_lines.extend([r"\bottomrule", r"\end{tabular}", r"\end{table}", ""])
    OUT_SNIPPET.parent.mkdir(parents=True, exist_ok=True)
    OUT_SNIPPET.write_text("\n".join(tex_lines), encoding="utf-8")


def main():
    rows = load_rows()
    cleaned, anomalies = clean_rows(rows)
    summary = summarize(cleaned)
    write_outputs(cleaned, anomalies, summary)
    print(f"Reviewed records: {summary['reviewed_records']}")
    print(f"Wrote {OUT_REPORT}")
    print(f"Wrote {OUT_JSON}")
    print(f"Wrote {OUT_CSV}")
    print(f"Wrote {OUT_SNIPPET}")


if __name__ == "__main__":
    main()
