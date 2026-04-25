import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
VAL_DIR = ROOT / "writing_outputs" / "20260412_autoteakg_silver_paper" / "data" / "validation_v4"
SAMPLE_CSV = VAL_DIR / "validation_sample_v4.csv"
WORKSHEET_CSV = VAL_DIR / "validation_worksheet_v4.csv"
OUT_XLSX = VAL_DIR / "AutoTeaKG_validation_v4_for_undergraduates.xlsx"
GUIDE_MD = VAL_DIR / "validation_guideline_v4_undergraduate_cn.md"


FIELD_EXPLANATIONS = [
    ("activity_category", "这条证据说的是哪类功能活性，如抗炎、抗氧化、代谢改善、肠道菌群调节等。"),
    ("study_type", "这篇研究属于哪种研究类型，如动物实验、随机对照试验、队列研究、系统综述等。"),
    ("evidence_level", "证据强度层级。动物实验一般是 preclinical_in_vivo，人群干预是 human_interventional。"),
    ("silver_component_group", "研究对象属于哪类茶成分，如 catechins、tea polysaccharides、mixed polyphenols。"),
    ("silver_processing_step", "茶材料是否报告了加工步骤，如发酵、干燥、烘焙、杀青等。没有报告就应为空或 not_applicable。"),
    ("silver_extraction_method", "茶材料是否报告了提取/制备方法，如水提、乙醇提取、超声提取、冲泡等。没有报告就应为空或 not_applicable。"),
    ("mechanism_label", "这条证据中的机制描述，如 SCFA、JAK2/STAT3、gut-liver axis、barrier function 等。"),
]

DECISIONS = ["correct", "minor_issue", "major_issue", "not_applicable"]
OVERALL = ["accept", "accept_with_correction", "reject"]
ISSUES = [
    "wrong_activity",
    "wrong_study_type",
    "wrong_evidence_level",
    "wrong_component",
    "wrong_processing",
    "wrong_extraction",
    "wrong_mechanism",
    "hallucinated_detail",
    "out_of_scope",
    "duplicate_or_unclear",
]


def read_csv(path):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def style_header(row):
    for cell in row:
        cell.font = Font(name="Arial", bold=True, color="1F3438")
        cell.fill = PatternFill("solid", fgColor="E8F2EE")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_all(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_validation(ws, col_name, formula, max_row):
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    col = headers[col_name]
    letter = get_column_letter(col)
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "请从下拉列表选择，不要手打其他值。"
    dv.prompt = "请选择标准标注值。"
    ws.add_data_validation(dv)
    dv.add(f"{letter}2:{letter}{max_row}")


def build_joined_rows(sample, worksheet):
    worksheet_by_id = {row["audit_id"]: row for row in worksheet}
    rows = []
    for s in sample:
        w = worksheet_by_id.get(s["audit_id"], {})
        row = {
            "audit_id": s.get("audit_id", ""),
            "record_id": s.get("record_id", ""),
            "paper_id": s.get("paper_id", ""),
            "title": s.get("title", ""),
            "abstract": s.get("abstract", ""),
            "claim_text_raw": s.get("claim_text_raw", ""),
            "activity_category_model": s.get("activity_category", ""),
            "study_type_model": s.get("study_type", ""),
            "evidence_level_model": s.get("evidence_level", ""),
            "component_group_model": s.get("silver_component_group", ""),
            "processing_step_model": s.get("silver_processing_step", ""),
            "extraction_method_model": s.get("silver_extraction_method", ""),
            "mechanism_label_model": s.get("mechanism_label", ""),
            "uncertainty_class": s.get("uncertainty_class", ""),
            "uncertainty_flags": s.get("uncertainty_flags", ""),
            "activity_category_decision": w.get("activity_category_decision", ""),
            "activity_category_corrected": w.get("activity_category_corrected", ""),
            "study_type_decision": w.get("study_type_decision", ""),
            "study_type_corrected": w.get("study_type_corrected", ""),
            "evidence_level_decision": w.get("evidence_level_decision", ""),
            "evidence_level_corrected": w.get("evidence_level_corrected", ""),
            "component_group_decision": w.get("silver_component_group_decision", ""),
            "component_group_corrected": w.get("silver_component_group_corrected", ""),
            "processing_step_decision": w.get("silver_processing_step_decision", ""),
            "processing_step_corrected": w.get("silver_processing_step_corrected", ""),
            "extraction_method_decision": w.get("silver_extraction_method_decision", ""),
            "extraction_method_corrected": w.get("silver_extraction_method_corrected", ""),
            "mechanism_label_decision": w.get("mechanism_label_decision", ""),
            "mechanism_label_corrected": w.get("mechanism_label_corrected", ""),
            "overall_record_decision": w.get("overall_record_decision", ""),
            "major_issue_type": w.get("major_issue_type", ""),
            "reviewer_id": w.get("reviewer_id", ""),
            "review_date": w.get("review_date", ""),
            "comments": w.get("comments", ""),
        }
        rows.append(row)
    return rows


def write_guide():
    guide = """# AutoTeaKG-Silver 48 条验证样本标注指南（本科生版）

## 1. 这项标注要做什么

你需要检查模型自动抽取的 48 条证据记录是否基本正确。每条记录来自一篇茶相关论文，Excel 里已经给出标题、摘要、模型抽取的字段，以及需要你填写的判断列。

请注意：你不是在判断“茶有没有效果”，而是在判断“模型有没有把论文里的信息抽对”。

## 2. 每条记录看哪些内容

优先看这几列：

1. `title`：论文标题。
2. `abstract`：论文摘要，是最重要的判断依据。
3. `claim_text_raw`：模型认为最关键的一句话。
4. 以 `_model` 结尾的列：模型抽取结果。
5. 以 `_decision` 结尾的列：你要填写是否正确。
6. 以 `_corrected` 结尾的列：如果模型错了，你填写更正值。

## 3. decision 怎么填

每个字段的 decision 只能填下面四种：

- `correct`：模型抽取基本正确。
- `minor_issue`：方向对，但不够精确。例如机制写得太宽泛，但没有乱编。
- `major_issue`：明显错误。例如动物实验被写成人体研究，或者机制是摘要里没有的。
- `not_applicable`：这篇论文没有报告这个字段，或这个字段不适用。

## 4. overall_record_decision 怎么填

- `accept`：整体可以接受，最多有很小问题。
- `accept_with_correction`：有问题，但通过更正字段可以保留。
- `reject`：这条记录整体不可靠，建议不要进入高质量分析。

## 5. 各字段怎么判断

### activity_category

看这条证据说的是哪类功能。常见类别：

- `anti-inflammatory`：抗炎、降低炎症因子、炎症损伤减轻。
- `antioxidant`：抗氧化、ROS、MDA、GSH、SOD 等。
- `gut microbiota modulation`：肠道菌群、Akkermansia、Lactobacillus、SCFAs 等。
- `anti-obesity`：体重、脂肪、肥胖、产热。
- `metabolic improvement`：血糖、胰岛素抵抗、NAFLD、脂质代谢。
- `neuroprotection`：认知、神经炎症、脑损伤。
- `cardiovascular protection`：心血管风险、动脉粥样硬化、死亡率。
- `other`：不属于上面类别，或模型没法归类。

### study_type

看研究类型：

- `animal study`：小鼠、大鼠等动物实验。
- `in vitro`：细胞或体外实验。
- `randomized controlled trial`：随机对照人体试验。
- `cohort study`：队列或观察性人群研究。
- `meta-analysis`：荟萃分析。
- `systematic review`：系统综述或普通综述。

### evidence_level

通常由 study_type 推出：

- 动物实验：`preclinical_in_vivo`
- 体外实验：`low_preclinical` 或 `in_vitro`
- 人体干预：`human_interventional`
- 人群观察：`human_observational`
- meta-analysis：`evidence_synthesis`
- review：`evidence_synthesis_nonquantitative`

### component_group

看研究对象是什么成分：

- EGCG、catechin：`catechins`
- tea polyphenols：`mixed polyphenols`
- tea polysaccharides、TPS：`tea polysaccharides`
- caffeine：`caffeine`
- theanine：`theanine`
- theaflavins：`theaflavins`
- tea extract：`whole extract`

### processing_step

只在论文明确提到茶材料加工时填写。例如：

- fermentation / fermented tea：发酵
- roasting：烘焙
- drying：干燥
- steaming / fixation：蒸青、杀青

如果摘要只是说 disease aging、middle-aged、freeze-dried sperm 这类不是茶加工，不要算 processing。

### extraction_method

只在论文明确提到提取或制备方法时填写。例如：

- water extract / aqueous extract：水提
- ethanol extraction：乙醇提取
- ultrasound extraction：超声提取
- infusion / brewing：冲泡
- isolation / purification：分离纯化

如果没有提到，就填 `not_applicable` 或留 corrected 空。

### mechanism_label

看模型写的机制是否在摘要里有依据。例如：

- SCFA-mediated JAK2/STAT3 signaling
- gut-liver axis
- barrier function
- NF-kB / MAPK / Nrf2 pathway

如果机制太宽泛但方向对，可以填 `minor_issue`。如果摘要里完全没有这个机制，就是 `major_issue`。

## 6. comments 怎么写

简单写清楚原因即可，例如：

- “摘要明确写了 mice，应为 animal study。”
- “摘要没有提取方法，extraction 应为 not_applicable。”
- “机制方向正确，但 JAK2/STAT3 只在摘要最后一句出现。”

## 7. 工作建议

每条记录建议 2-4 分钟。先看标题和 claim_text_raw，再看摘要中相关句子。不要查太多外部资料，除非摘要确实看不懂。
"""
    GUIDE_MD.write_text(guide, encoding="utf-8")


def main():
    sample = read_csv(SAMPLE_CSV)
    worksheet = read_csv(WORKSHEET_CSV)
    rows = build_joined_rows(sample, worksheet)
    wb = Workbook()
    ws_intro = wb.active
    ws_intro.title = "使用说明"
    intro_lines = [
        ("目标", "检查模型抽取的 48 条茶功能活性证据是否正确。"),
        ("判断依据", "优先看 title、abstract 和 claim_text_raw。"),
        ("填写方式", "在黄色 decision / corrected / overall / comments 列填写。"),
        ("不要做什么", "不要判断茶是否真的有效；只判断模型是否抽对论文信息。"),
        ("建议速度", "每条 2-4 分钟，先做完一轮，再回头处理难例。"),
    ]
    for r, (k, v) in enumerate(intro_lines, start=1):
        ws_intro.cell(r, 1, k)
        ws_intro.cell(r, 2, v)
    ws_intro.column_dimensions["A"].width = 18
    ws_intro.column_dimensions["B"].width = 90
    style_all(ws_intro)

    ws = wb.create_sheet("标注主表")
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    style_header(ws[1])
    style_all(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    width_map = {
        "audit_id": 10,
        "record_id": 22,
        "paper_id": 16,
        "title": 46,
        "abstract": 88,
        "claim_text_raw": 48,
        "mechanism_label_model": 32,
        "uncertainty_flags": 36,
        "comments": 42,
    }
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width_map.get(h, 18)
        if h.endswith("_decision") or h.endswith("_corrected") or h in {"overall_record_decision", "major_issue_type", "reviewer_id", "review_date", "comments"}:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row_idx, i).fill = PatternFill("solid", fgColor="FFF2CC")
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 92

    ws_fields = wb.create_sheet("字段解释")
    ws_fields.append(["字段", "怎么理解"])
    for row in FIELD_EXPLANATIONS:
        ws_fields.append(list(row))
    style_header(ws_fields[1])
    style_all(ws_fields)
    ws_fields.column_dimensions["A"].width = 28
    ws_fields.column_dimensions["B"].width = 100

    ws_lists = wb.create_sheet("下拉选项")
    ws_lists.append(["decision", "overall", "major_issue_type"])
    max_len = max(len(DECISIONS), len(OVERALL), len(ISSUES))
    for i in range(max_len):
        ws_lists.append([
            DECISIONS[i] if i < len(DECISIONS) else "",
            OVERALL[i] if i < len(OVERALL) else "",
            ISSUES[i] if i < len(ISSUES) else "",
        ])
    style_header(ws_lists[1])
    style_all(ws_lists)
    ws_lists.sheet_state = "hidden"

    for h in headers:
        if h.endswith("_decision"):
            add_validation(ws, h, "=下拉选项!$A$2:$A$5", ws.max_row)
    add_validation(ws, "overall_record_decision", "=下拉选项!$B$2:$B$4", ws.max_row)
    add_validation(ws, "major_issue_type", "=下拉选项!$C$2:$C$11", ws.max_row)

    wb.save(OUT_XLSX)
    write_guide()
    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {GUIDE_MD}")


if __name__ == "__main__":
    main()
