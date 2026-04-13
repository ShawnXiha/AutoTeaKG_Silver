import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
TEMPLATES_DIR = ROOT / "templates"
PROTOCOLS_DIR = ROOT / "protocols"

SAMPLE_CSV = TEMPLATES_DIR / "llm_manual_audit_sample_2026-04-02.csv"
WORKSHEET_CSV = TEMPLATES_DIR / "llm_manual_audit_worksheet_2026-04-02.csv"
GUIDELINE_MD = PROTOCOLS_DIR / "llm_manual_audit_guideline_2026-04-02.md"
OUT_XLSX = TEMPLATES_DIR / "llm_manual_audit_workbook_2026-04-03.xlsx"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
INFO_FILL = PatternFill("solid", fgColor="F3F6FA")


def read_csv(path: Path):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def style_header(row):
    for cell in row:
        cell.font = Font(name=FONT_NAME, bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_sheet_defaults(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name=FONT_NAME, size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def set_widths(ws, width_map, default=16):
    max_col = ws.max_column
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        header = ws.cell(1, col_idx).value or ""
        ws.column_dimensions[col_letter].width = width_map.get(header, default)


def write_table_sheet(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    style_header(ws[1])
    style_sheet_defaults(ws)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def build_joined_rows(sample_rows, worksheet_rows):
    worksheet_by_audit = {row["audit_id"]: row for row in worksheet_rows}
    joined = []
    for sample in sample_rows:
        worksheet = worksheet_by_audit.get(sample["audit_id"], {})
        joined.append(
            {
                "audit_id": sample.get("audit_id", ""),
                "paper_id": sample.get("paper_id", ""),
                "record_id": sample.get("record_id", ""),
                "title": sample.get("title", ""),
                "abstract": sample.get("abstract", ""),
                "abstract_source": sample.get("abstract_source", ""),
                "journal": sample.get("journal", ""),
                "year": sample.get("year", ""),
                "sampling_reason": sample.get("sampling_reason", ""),
                "annotator_id": sample.get("annotator_id", ""),
                "confidence_score": sample.get("confidence_score", ""),
                "activity_category_pred": sample.get("activity_category", ""),
                "endpoint_label_pred": sample.get("endpoint_label", ""),
                "study_type_pred": sample.get("study_type", ""),
                "evidence_level_pred": sample.get("evidence_level", ""),
                "effect_direction_pred": sample.get("effect_direction", ""),
                "mechanism_label_pred": sample.get("mechanism_label", ""),
                "microbiota_taxon_pred": sample.get("microbiota_taxon", ""),
                "microbial_metabolite_pred": sample.get("microbial_metabolite", ""),
                "host_phenotype_pred": sample.get("host_phenotype", ""),
                "claim_text_raw": sample.get("claim_text_raw", ""),
                "field_activity_category": worksheet.get("field_activity_category", ""),
                "field_endpoint_label": worksheet.get("field_endpoint_label", ""),
                "field_study_type": worksheet.get("field_study_type", ""),
                "field_evidence_level": worksheet.get("field_evidence_level", ""),
                "field_effect_direction": worksheet.get("field_effect_direction", ""),
                "field_mechanism_label": worksheet.get("field_mechanism_label", ""),
                "field_microbiota_taxon": worksheet.get("field_microbiota_taxon", ""),
                "field_microbial_metabolite": worksheet.get("field_microbial_metabolite", ""),
                "field_host_phenotype": worksheet.get("field_host_phenotype", ""),
                "paper_scope_decision": worksheet.get("paper_scope_decision", ""),
                "overall_record_decision": worksheet.get("overall_record_decision", ""),
                "major_issue_type": worksheet.get("major_issue_type", ""),
                "corrected_activity_category": worksheet.get("corrected_activity_category", ""),
                "corrected_endpoint_label": worksheet.get("corrected_endpoint_label", ""),
                "corrected_study_type": worksheet.get("corrected_study_type", ""),
                "corrected_evidence_level": worksheet.get("corrected_evidence_level", ""),
                "corrected_effect_direction": worksheet.get("corrected_effect_direction", ""),
                "corrected_mechanism_label": worksheet.get("corrected_mechanism_label", ""),
                "corrected_microbiota_taxon": worksheet.get("corrected_microbiota_taxon", ""),
                "corrected_microbial_metabolite": worksheet.get("corrected_microbial_metabolite", ""),
                "corrected_host_phenotype": worksheet.get("corrected_host_phenotype", ""),
                "reviewer_id": worksheet.get("reviewer_id", ""),
                "review_date": worksheet.get("review_date", ""),
                "comments": worksheet.get("comments", ""),
            }
        )
    return joined


def add_validation_lists(ws):
    lists = {
        "field_status": ["correct", "minor_issue", "major_issue", "not_applicable"],
        "scope_decision": ["in_scope", "borderline", "out_of_scope"],
        "record_decision": ["accept", "accept_with_correction", "reject"],
        "issue_type": [
            "wrong_activity",
            "wrong_endpoint",
            "wrong_study_type",
            "wrong_evidence_level",
            "wrong_direction",
            "hallucinated_mechanism",
            "hallucinated_taxon",
            "out_of_scope_paper",
            "duplicate_record",
        ],
    }
    ws.title = "Lists"
    for col_idx, (name, values) in enumerate(lists.items(), start=1):
        ws.cell(1, col_idx, name)
        for row_idx, value in enumerate(values, start=2):
            ws.cell(row_idx, col_idx, value)
    style_header(ws[1])
    style_sheet_defaults(ws)
    return {
        "field_status": "=Lists!$A$2:$A$5",
        "scope_decision": "=Lists!$B$2:$B$4",
        "record_decision": "=Lists!$C$2:$C$4",
        "issue_type": "=Lists!$D$2:$D$10",
    }


def apply_validations(ws, refs):
    field_columns = [
        "field_activity_category",
        "field_endpoint_label",
        "field_study_type",
        "field_evidence_level",
        "field_effect_direction",
        "field_mechanism_label",
        "field_microbiota_taxon",
        "field_microbial_metabolite",
        "field_host_phenotype",
    ]
    header_to_col = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}

    def add_list_validation(column_name, formula):
        col_idx = header_to_col[column_name]
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Please choose a value from the dropdown list."
        dv.prompt = "Select a standardized audit value."
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{ws.max_row}")

    for name in field_columns:
        add_list_validation(name, refs["field_status"])
    add_list_validation("paper_scope_decision", refs["scope_decision"])
    add_list_validation("overall_record_decision", refs["record_decision"])
    add_list_validation("major_issue_type", refs["issue_type"])


def highlight_input_columns(ws):
    input_headers = {
        "field_activity_category",
        "field_endpoint_label",
        "field_study_type",
        "field_evidence_level",
        "field_effect_direction",
        "field_mechanism_label",
        "field_microbiota_taxon",
        "field_microbial_metabolite",
        "field_host_phenotype",
        "paper_scope_decision",
        "overall_record_decision",
        "major_issue_type",
        "corrected_activity_category",
        "corrected_endpoint_label",
        "corrected_study_type",
        "corrected_evidence_level",
        "corrected_effect_direction",
        "corrected_mechanism_label",
        "corrected_microbiota_taxon",
        "corrected_microbial_metabolite",
        "corrected_host_phenotype",
        "reviewer_id",
        "review_date",
        "comments",
    }
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(1, col_idx).value
        if header in input_headers:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row_idx, col_idx).fill = INPUT_FILL


def build_overview_sheet(ws, sample_rows):
    llm_rows = [row for row in sample_rows if row.get("annotator_id") == "glm5_nvidia"]
    low_conf = 0
    microbiome = 0
    review = 0
    for row in sample_rows:
        try:
            confidence = float(row.get("confidence_score") or 0)
        except ValueError:
            confidence = 0.0
        if confidence < 0.80:
            low_conf += 1
        if row.get("activity_category") == "gut microbiota modulation" or row.get("microbiota_taxon") or row.get("microbial_metabolite"):
            microbiome += 1
        if row.get("study_type") in {"systematic review", "meta-analysis"}:
            review += 1

    lines = [
        ("Workbook", "LLM Manual Audit Workbook"),
        ("Date", "2026-04-03"),
        ("Sample records", str(len(sample_rows))),
        ("LLM records", str(len(llm_rows))),
        ("Low-confidence records", str(low_conf)),
        ("Microbiome-related records", str(microbiome)),
        ("Review/meta-analysis records", str(review)),
        ("How to use", "Open Audit_Form, read title + abstract + predicted fields, then fill the yellow reviewer columns."),
        ("Decision rule", "Use accept / accept_with_correction / reject for overall_record_decision."),
        ("Note", "Sample_Raw and Worksheet_Template are preserved as raw exports; Guideline contains the full audit instructions."),
    ]
    for row_idx, (label, value) in enumerate(lines, start=1):
        ws.cell(row_idx, 1, label)
        ws.cell(row_idx, 2, value)
    ws["A1"].font = Font(name=FONT_NAME, bold=True, size=12)
    for row_idx in range(1, len(lines) + 1):
        ws.cell(row_idx, 1).font = Font(name=FONT_NAME, bold=True)
        ws.cell(row_idx, 1).fill = INFO_FILL
        ws.cell(row_idx, 2).font = Font(name=FONT_NAME)
        ws.cell(row_idx, 2).alignment = Alignment(vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 120
    ws.freeze_panes = "A1"


def build_guideline_sheet(ws, guideline_text):
    for row_idx, line in enumerate(guideline_text.splitlines(), start=1):
        ws.cell(row_idx, 1, line)
    ws.column_dimensions["A"].width = 150
    style_sheet_defaults(ws)


def main():
    sample_rows = read_csv(SAMPLE_CSV)
    worksheet_rows = read_csv(WORKSHEET_CSV)
    guideline_text = GUIDELINE_MD.read_text(encoding="utf-8")

    wb = Workbook()
    overview_ws = wb.active
    overview_ws.title = "Overview"
    build_overview_sheet(overview_ws, sample_rows)

    audit_ws = wb.create_sheet("Audit_Form")
    joined_rows = build_joined_rows(sample_rows, worksheet_rows)
    audit_headers = list(joined_rows[0].keys()) if joined_rows else []
    write_table_sheet(audit_ws, audit_headers, joined_rows)
    audit_ws.freeze_panes = "D2"
    audit_ws.sheet_view.zoomScale = 85
    width_map = {
        "audit_id": 12,
        "paper_id": 16,
        "record_id": 22,
        "title": 42,
        "abstract": 90,
        "abstract_source": 34,
        "journal": 24,
        "year": 10,
        "sampling_reason": 22,
        "annotator_id": 16,
        "confidence_score": 12,
        "claim_text_raw": 44,
        "comments": 36,
    }
    set_widths(audit_ws, width_map, default=18)
    for row_idx in range(2, audit_ws.max_row + 1):
        audit_ws.row_dimensions[row_idx].height = 72
    highlight_input_columns(audit_ws)

    sample_ws = wb.create_sheet("Sample_Raw")
    write_table_sheet(sample_ws, list(sample_rows[0].keys()), sample_rows)
    set_widths(sample_ws, {"title": 42, "abstract": 90, "abstract_source": 34, "claim_text_raw": 44}, default=18)
    for row_idx in range(2, sample_ws.max_row + 1):
        sample_ws.row_dimensions[row_idx].height = 72

    template_ws = wb.create_sheet("Worksheet_Template")
    write_table_sheet(template_ws, list(worksheet_rows[0].keys()), worksheet_rows)
    set_widths(template_ws, {"comments": 36}, default=18)

    guideline_ws = wb.create_sheet("Guideline")
    build_guideline_sheet(guideline_ws, guideline_text)

    refs = add_validation_lists(wb.create_sheet())
    wb["Lists"].sheet_state = "hidden"
    apply_validations(audit_ws, refs)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
